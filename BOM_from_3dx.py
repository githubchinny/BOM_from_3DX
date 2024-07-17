# %% [markdown]
# # Build a new BoM from 3dx files
# 
# This is reading in extracts produced by 3dx, adding the function group and system/subsystem, and doing the roll up quantity calcs in the absence of 3DX doing that.
# 
# The script does the following functions that is missing from 3DX currently:
# 
# * Populate Function Group/System/Sub System – not fully populated in 3DX for all parts and this is needed downstream for reporting
# * Add Percentage Missing calculations for all attributes/columns for EDAG reporting
# * Create a matching key for Smartsheets processing
# * Add last Export Date to top of export and as a column 
# * Add BOM COUNT validation column
# * Add Parent Part column – not maintained in 3DX and requested by downstream systems
# * Validates all mandatory attributes are included before processing any further
# * If Quantity doesn’t exist (comes only from XEN tool) then group parts and create Quantity column (removes the reliance on XEN license)
# * Clear zero values, ‘not set’ and blank spaces from 3DX extract – zero does not mean the attribute is populated and needs to be removed to show it requires completion
# * Add Part Number validation and report where it doesn’t adhere to part numbering standard – creates validation columns at the far right of the export
# * Force the order of selected key attributes to the left as not always maintained in 3DX extract files
# * Derive the export folder to send file to and name file for Data Shuttle to pick up – top level ‘Title’ (ie T48e-01-Z00001 = VP 5 door master product - T48e-01-Z00001’
# * Identify where a part is a lowest level child part and add a ‘has_child’ column to support Smartsheet Mass calculations
# 

# %%
import pandas as pd
import numpy as np
import os
import sys
import re
import io
import time
import openpyxl
from openpyxl import load_workbook
import xlwings as xw
import glob
import configparser
from pathlib import Path
import datetime
import platform


# %%
def find_files(download_dir):
    # find any changed files changed in past 2hrs in the downloads directory
    dirpath = download_dir
    past = time.time() - 2*60*60 # 2 hours
    files = []
    for p, ds, fs in os.walk(dirpath):
        for fn in fs:
            # was using this to filter what filenames to find
            if 'ENO' in fn:            
                filepath = os.path.join(p, fn)
                if os.path.getmtime(filepath) >= past:
                    files.append(filepath)

    return files

# %%
def open_file(file):
    # get 3dx extract
    print ("opening {} at: {}".format(file, time.strftime("%H:%M:%S", time.localtime())))
    with open(file, "rb") as f:
        try:

            if 'csv' in file:
                BOM = pd.DataFrame()
                BOM = pd.read_csv(f, low_memory=False, na_values="", keep_default_na=False) 
                # skip first n rows that are header information out of 3DX
                if 'Level' not in BOM.columns:
                    print ("didn't get the column headers so finding them myself")            
                    n = BOM[BOM.iloc[:, 0] == 'Level'].index.values[0]
                    # create the column headers
                    BOM.columns = BOM.iloc[n]
                    # drop the top rows above the header row we found
                    BOM = BOM[n+1:]

            if 'xls' in file:
                # BOM = pd.read_csv(f, low_memory=False, skiprows=8) 
                BOM = pd.read_excel(f)             
            # sheetnames = [sheet for sheet in f.sheet_names]
        except Exception as e:
            print ("{}".format(e))

    return BOM


# %%
def add_function_group(df):
    # Add Function and Sub Group if it doesn't already exist

    # - Level 0 = Model Variant   
    # - Level 1 = Function Group Area   
    # - Level 2 = System   
    # - Level 3 = Sub Systems
    # - level 4 = AMs/SAs??

    # Find each one and forward fill to the next occurrence
    # function group - level 1
    df['Function Group'] = np.where(df['Level'].isin([0,1]), df['Description'], None)
    df['Function Group'] = df['Function Group'].ffill()

    # System - level 2
    df['System'] = np.where(df['Level'] == 2, df['Description'], None)
    df['System'] = np.where(df['Level'] >= 2, df['System'].ffill(), None)

    # SUB_System = level 3
    df['Sub System'] = np.where(df['Level'] == 3, df['Description'], None)
    df['Sub System'] = np.where(df['Level'] >= 3, df['Sub System'].ffill(), df['Sub System'])

    # Level 4 - for grouping mass and cost roll up for Carlos/smartsheets
    df['Level_4_Parent'] = np.where(df['Level'] == 4, df['Title'], None)
    df['Level_4_Parent'] = np.where(df['Level'] >= 4, df['Level_4_Parent'].ffill(), df['Level_4_Parent'])

    return df

# %%
def rename_columns(df):
    # keep python index as 'orig_sort' - useful for knowing you've maintained the BoM structure/order
    df.reset_index(inplace=True)
    df.rename(columns={'index':'orig_sort'}, inplace=True)

    # sometimes column came through as Title (Instance) 
    try:
        df.rename(columns={'Title (Instance)':'Instance Title'}, inplace=True)
    except ValueError:
        print ("didn't find Title (Instance) to rename")

    # 3dx calls Quantity Occurrences
    try:
        df.rename(columns={'Occurrences':'Quantity'}, inplace=True)
    except ValueError:
        print ("didn't find Occurrences to rename")

    # replace any Mass (kg) columns with Mass
    df.columns = df.columns.str.replace('Mass (kg)', 'Mass')


    return df  

# %%
def add_bi_key(df):
    # for use in power_bi reporting
    # replace NaN with ''
    df['bi_combined_key'] = df['Variant'].astype(str) + df['Function Group'].astype(str) + df['System'].astype(str) + df['Sub System'].astype(str)
    
    return df['bi_combined_key']

# %%
def add_matching_key(df):
    # for us in smartsheets
    df['Matching Key'] = np.where(df['Parent Part'].isna(), df['Title'].astype(str), df["Title"].astype(str) + df["Parent Part"].astype(str))
    # force matching key to always be upper case whilst there is inconsistency within 3dx.  Otherwise, doesn't match dup parts
    df['Matching Key'] = df['Matching Key'].str.upper()
    # build cumulative count for each part
    df['cumcount'] = df.groupby('Matching Key').cumcount()+1
    # not interested in the first occurrence of a part - blank out the first cumcount of each group
    df['cumcount'] = np.where(df['cumcount']==1, '', df['cumcount'])
    df['Matching Key'] = df['Matching Key'].astype(str) + df['cumcount'].astype(str)

    return df

# %%
def add_smartsheet_cols(df, extract_date):
    # need to correct the percent missing for matching key column
    df.loc['percent_missing','Matching Key'] = 0
    df['Last Export Date'] = extract_date

    return df

# %%
def mandatory_attributes(bom_cols):
    # read in mandatory fields from file in same directory
    mand_file = Path(sys.path[0]) / 'mandatory_attributes.ini'
    with open(mand_file, 'r') as f:
        lst = f.readlines()

    mand_cols = [line.rstrip() for line in lst]

    missing_cols = list(set(mand_cols) - set(bom_cols))

    if len(missing_cols) > 0:
        print ("missing mandatory attributes: {}".format(missing_cols))
        sys.exit()

    return mand_cols

# %%
def order_columns(df):
    # this specifies the order of the left most cols.  Columns not mentioned below then appear alphabetically afterwards
    cols_to_order = ['BOM COUNT',
        'Matching Key',
        'Last Export Date',
        'orig_sort',
        'Function Group',
        'System',
        'Sub System',
        'Level_4_Parent',
        'Level',
        'Title',
        'Parent Part',
        'Revision',
        'Description',
        'Name',
        'Quantity',
        'Source Code',
        'UOM',
        'Provide',
        'CAD Mass',
        'CAD Material']
        # 'Subtype']
    
    # # write out mandatory attributes to local file.  Not reading in from here as there is always some code change required anyway!
    # with open('mandatory_attributes.ini', 'w') as f:
    #     for col in cols_to_order:
    #         # ignore the internal cols I create
    #         if col not in (['Matching Key','Last Export Date','orig_sort','Parent Part']):
    #             f.write('{}\n'.format(col))

    try:
        ordered_cols = cols_to_order + (df.columns.sort_values().drop(cols_to_order).tolist())
        df = df[ordered_cols]
    except KeyError as e:
        raise Exception ("Missing an expected column in the extract: {}".format(e))

    return df

# %%
def create_sa_index(df):
    # fill level 3s with orig_sort
    # ffill everything with the level 3 orig_sort
    # fill level 4s with the level 3 orig_sort + its own orig_sort
    # fill < level 3 with own orig_sort
    # NaN > level 4 and refill with the sa_index from level 4 above
    df['SA_Index'] = np.where(df['Level'] == 3, df['orig_sort'].astype(str), np.nan)
    df['SA_Index'] = df['SA_Index'].ffill()
    df['SA_Index'] = np.where(df['Level'] == 4, df['SA_Index'] + '_' + df['orig_sort'].astype(str), df['SA_Index'])
    # forward fill so that > Level 5 get the same index
    df['SA_Index'] = np.where(df['Level'] < 3, df['orig_sort'].astype(str), df['SA_Index'])
    df['SA_Index'] = np.where(df['Level'] > 4, np.nan, df['SA_Index'])
    df['SA_Index'] = df['SA_Index'].ffill()


    return df

# %%
def create_sa_index2(df):
    # this was attempting to rely on 3dx labelling of an Assembly - didn't work all of the time
    df['SA_Index'] = np.where(df['Assembly'], df['orig_sort'].astype(str), np.nan)
    df['SA_Index'] = df['SA_Index'].ffill()

    return df

# %%
def add_quantities(df):

    # if the extract hasn't come with quantity column then we need to calculate our own grouping of like parts together
    # this is only being done for level 4 and greater - level 4 is assumed to be assembly level
    # there is a potential problem with this if the extract is not produced for the top level part of a structure
    
    # groupby:
    # SA_Index - created earlier, this will group the parts within an assembly (level 4 and greater) 
    # Title - this is the part number
    # Parent Part - created earlier, this will group only parts at the same level with the same parent
    # Level - this is probably not required if we are using parent part, but ensures we group at the same level

    # groupby Title and sum (size).  Save as a new df called qty
    qty = BOM_pp.groupby(['SA_Index','Title','Parent Part','Level'], dropna=False).size().reset_index(name='Quantity')
    # qty = BOM_pp.groupby(['Title','Parent Part','Level'], dropna=False).size().reset_index(name='Quantity')

    # merge qty with BOM on SA_Index to get all the other columns back
    qty2 = pd.merge(qty, BOM_pp, on=['SA_Index','Title','Parent Part','Level'])
    # qty2 = pd.merge(qty, BOM_pp, on=['Title','Parent Part','Level'])
    # qty2 = pd.concat([qty, BOM_pp])

    # need to drop dups using only a subset of cols, creating new_bom df
    new_bom = qty2.drop_duplicates(subset=['SA_Index','Title','Parent Part', 'Level'])
    # new_bom = qty2.drop_duplicates(subset=['Title','Parent Part', 'Level'])
    # sort the new_nom df by the orig_sort field to make sure it's the order it came out of 3dx
    new_bom = new_bom.sort_values(by='orig_sort')
    
    # don't think there are any names to rename?
    new_bom.rename(columns={
        'Title_y':'Title',
        'Parent Part_x':'Parent Part',
        'Quantity_x':'Quantity',
        'Level_x':'Level'
    }, inplace=True)    

    return new_bom


# %%
def create_parent_part(df):
    # reset index before trying to update, otherwise multiple rows get updated
    df.reset_index(inplace=True, drop=True)
    
    df['Parent Part'] = None
    df['Parent Revision'] = None

    level = {}
    previous_parent_part=0

    for i, row in df.iterrows():
        current_part_number = row[['Title','Revision']]
        current_part_level = row['Level']


        # write part number and revision to dictionary under current part level
        level[current_part_level] = current_part_number

        # reset higher levels for each assembly
        # remove entries from higher levels
        keys = [k for k in level if k > current_part_level]
        for x in keys:
            del level[x]

        if current_part_level > 0:
            # get the max part level from the level dictionary that's less than current part level
            previous_parent_level = max(k for k in level if k < current_part_level)

            # update the parent part and parent part revision
            # print (i, "Parent part {} from previous level {}".format(level[previous_parent_level], previous_parent_level))
            df.at[i,'Parent Part'] = level[previous_parent_level].iloc[0]
            df.at[i,'Parent Revision'] = level[previous_parent_level].iloc[1]
            
    return df

# %%
def create_gparent_part(df):
    # reset index before trying to update, otherwise multiple rows get updated
    df.reset_index(inplace=True, drop=True)
    
    df['Parent Part'] = None
    df['Matching Key'] = None

    level = {}
    previous_part_level=0
    gparent_part_level=0

    for i, row in df.iterrows():
        current_part_number = row['Title']
        current_part_level = row['Level']

        # write part number to dictionary under current part level
        level[current_part_level] = current_part_number

        # reset higher levels for each assembly
        # remove entries from higher levels
        keys = [k for k in level if k > current_part_level]
        for x in keys:
            del level[x]

        if current_part_level > 0:
            # get the max part level from the level dictionary that's less than current part level
            previous_part_level = max(k for k in level if k < current_part_level)

            # update the parent part
            df.at[i,'Parent Part'] = level[previous_part_level]
        
        # if previous_part_level > 0:
            # get the max part level from the level dictionary that's less than previous parent level
            # gparent_part_level = max(k for k in level if k < previous_part_level)
            # gparent_part_level = ''.join((level.values()))


            # update the parent part
            # print (i, "Parent part {} from previous level {}".format(level[previous_parent_level], previous_parent_level))
        df.at[i,'Matching Key'] = ''.join((level.values()))


    return df

# %%
def convert_to_xml(df):
    BOM_xml = df
    # get rid of spaces, slashes and chars xml can't handle
    BOM_xml.columns = df.columns.str.replace(' ', '_')
    BOM_xml.columns = df.columns.str.replace('/', '_')
    BOM_xml.columns = df.columns.str.replace('&', '')

    return df

# %%
def data_shuttle_folder(product):
    # After update of BoM names
    # •	VP 5 door master product - T48e-01-Z00001 
    # •	VP 3 door master product - T48e-02-Z00001
    # •	XP 5 door master product – T48e-01-Z00005
    # •	VP 3 door master product - T48e-02-Z00005

    config = configparser.ConfigParser()
    config_file = Path(sys.path[0]) / 'product_structures_config.ini'
    config.read(config_file)

    # read the Product_Structures section and look up the product
    try:
        ds_folder = config['Product_Structures'][product]
    except KeyError:
        print ("No entry in file {}, section: Product_Structures, for key {}.  Using 'default'".format(config_file, product))
        ds_folder = 'default'

    return ds_folder



# %%
def COG_split(df):
    # split the COG field (where it is populated) into COG x, COG y, COG z 
    # need to make sure COG is a string field - if is it completely empty, all nan will make it float
    df.COG = df.COG.astype('object')

    # then try to split into three columns
    try:
        df[['COG X', 'COG Y', 'COG Z']] = df['COG'].str.split(',', expand=True)
    except ValueError:
        # will fail if there is nothing to split on, so create 3 cols with NaN
        df[['COG X', 'COG Y', 'COG Z']] = np.nan

    return df

# %%
def clear_zero_values(df):
    value_cols = df.select_dtypes(exclude=[object]).columns
    value_cols = value_cols.drop(['Level','orig_sort'])

    # convert 0.0 to na
    df[value_cols] = np.where(df[value_cols] == 0, np.nan, df[value_cols])

    # for COG, convert 0,0,0 to na
    df['COG'] = np.where(df['COG'] == '0,0,0', np.nan, df['COG'])

    return df

# %%
def clear_not_set_values(df):
    df = df.replace('Not Set', np.nan)

    return df

# %%
def set_na_values(df):
    # populate Provide with 'N/A' for source code SYS and ENG
    df['Provide'] = np.where(df['Source Code'].isin(['SYS','ENG']), 'N/A', df['Provide'])

    return df

# %%
def percent_missing(df):
    df.loc['percent_missing'] = None
    df.loc['percent_missing'] = df.isnull().sum(axis=0) * 100 / len(df)
    # df.loc['percent_missing'] = df.loc['percent_missing'].astype(int)
    df.loc['percent_missing','orig_sort'] = 'percent_missing'
    new_df = pd.concat([df.iloc[-1:].copy(), df.iloc[:-1].copy()])

    return new_df

# %%
def validate_part_no(df):

    # ? means the preceding bracketed group is optional (optional s,S and trailing X)
    # pattern = r'([A-Z]\d{2}([s,S])?-[A-Z]\d{4}(X)?)'
    # pattern = r'[A-Z]\d{2}[e]-[A-Z]\d{5}+X?'
    # pattern = r'([A-Z]\d{2}[e])-([A-Z])(\d{5})(X)'
    pattern = r'([A-Z]\d{2}[e])-(\w[A-Za-z0-9]*)?-?([A-Z])(\d{5})(X)?'
    df[['extr_project','extr_invalid_code','extr_function','extr_pn','extr_maturity']] = df['Title'].str.extract(pattern, expand=True)

    # done outside of this function now
    # df['part_number_length'] = df['Title'].str.len()

    return df


# %%
def clear_blanks(df):
    # replace an empty string and records with only spaces
    df = df.replace(r'^\s*$', np.nan, regex=True)

    return df

# %%
def convert_date(x):
    try:
        return dt.datetime.strptime(x, '%Y-%m-%d')
    except:
        return pd.NaT

# %%
def write_to_excel(df, outfile):
    from openpyxl import load_workbook

    # update if exists
    try:
        wb = load_workbook(outfile)
        with pd.ExcelWriter(outfile, engine='openpyxl') as writer:
            writer.workbook = wb
            df.to_excel(writer, index=False)
            print ("BOM written to existing file {}".format(outfile))
    except (FileNotFoundError):
        df.to_excel(outfile, index=False)
        print ("BOM written to new file {}".format(outfile))


# %% [markdown]
# XCADEmbeddedCmp means it is an Embedded Component, which means its behaves like an Assembly (so it has child parts), but it usually only exists in the context of its parent. So for your purposes, it should be flagged as an Assembly. 
# 
# I think KPKV5EquivalentComputed is just another element of the inertia measure function which is on the part template, so it appears whether or not its had the mass calculated. For the two parts that don't have KPKV5EquivalentComputed, it might mean they were created from a different template. 
# 
# ‘XCAD Extension’ and ‘XCADExtension’ are the same, I think its just displaying differently on its own than when its concatenated into a string with other values. If they don't have 3DPart then they are Assemblies. 
# 

# %%
def split_subtype(df):
    
    df.Subtype.replace({'XCAD Extension':'Assembly',
                        'XCADExtension':'Assembly',
                        'XCADEmbeddedCmp':'Assembly',
                            '3DPart':'Part3D',
                            'XCADExposedPLMParameterSet':'COG and Mass Calculated'}, 
                            regex=True,
                            inplace=True)
                
    df.Subtype = df.Subtype.str.split(',')
    df.Subtype = df.Subtype.fillna("").apply(list)
    # dynamically create columns
    for i in sorted(set(sum(df.Subtype.tolist(),[]))):
        # Create a new column 
        df[i] = df.Subtype.apply(lambda x: 1 if i in x else 0)

    return df

# %%
def has_child(df):
    df['has_child'] = np.where(df.Level>=df.Level.shift(-1), 0, 1)
    # set last row to has_child = 0 because there isn't anything below it
    df.loc[df.index[-1],'has_child'] = 0

    return df

# %%
def create_power_bi_df(df, product):
    # copy df without the 1st row of metrics and 1st col of BOM COUNTs
    temp = df.iloc[1:,1:].copy()
    for col in ['UOM','Provide','Source Code']:
        temp['Missing {}'.format(col)] = np.where(temp[col].isnull(), 1, 0)

    # add variant column for merging multiple BOMs together and reporting on 1 dashboard
    temp['Variant'] = product
    temp['bi_combined_key'] = add_bi_key(temp)

    return temp

# %%
def build_reports(df):

    # write all reports to same dict - will create individual sheets for each report later
    fg_reports = {}

    for col in ['Source Code','Provide','UOM']:
        # iloc -1 drops the bottom row of the dataframe, which will be the column totals if margins_names=True
        fg_reports['{}_by_FG'.format(col)] = pd.crosstab([df['Variant'],df['Last Export Date'],df['Function Group']], df[col], margins=True, margins_name='Totals', dropna=False).reset_index().iloc[:-1]
        fg_reports['{}_by_FG'.format(col)].rename(columns={np.nan:'Missing'}, inplace=True)
        # fg_reports['{} %ages by FG'.format(col)] = pd.crosstab(power_bi['Function Group'], power_bi[col], margins=True, margins_name='Totals', dropna=False, normalize=True).iloc[:-1].round(4)*100
        # fg_reports['{} %ages by FG'.format(col)].rename(columns={np.nan:'Missing'}, inplace=True)

    for col in ['Source Code','Provide','UOM']:
        # fg_combined['Missing {} %ages by FG, System, Sub System'.format(col)] = pd.crosstab([power_bi['Function Group'],power_bi['System'], power_bi['Sub System'], power_bi['bi_combined_key']], power_bi[col].isna(), margins=True, margins_name='Totals', normalize='index').iloc[:-1].round(4)*100
        # fg_combined['Missing {} %ages by FG, System, Sub System'.format(col)].rename(columns={False:'Populated', True:'Missing'}, inplace=True)
        fg_reports['Missing_{}'.format(col)] = pd.crosstab([df['Variant'],df['Last Export Date'],df['Function Group'],df['System'], df['Sub System'], df['bi_combined_key']], df[col].isna(), margins=True, margins_name='Totals').reset_index().iloc[:-1]
        fg_reports['Missing_{}'.format(col)].rename(columns={False:'Populated', True:'Missing'}, inplace=True)


    return fg_reports

# %%
def CAD_Material_validation(df):
    df[1:][df['Title'].str.contains('TPP', na=False)].groupby(['Title','CAD Material']).size()


# %%
def duplicate_checks(df):
# columns to use

    mask = ['Title', 'Revision', 'Description', 'Name', 'Source Code', 
        'UOM', 'Provide', 'Actual Mass', 'CAD Mass',
        'CAD Material', 'Programme Maturity', 'CAD Maturity',
        'CAD Surface Treatment', 'CAE Responsible',
        'Colour Relevant',
        'Electrical Connector', 'Estimated Mass', 'Evolution',
        'External Description', 'External Part Number', 'External Revision',
        'Part Identification', 'Part Quality Class', 'Part Thickness',
        'Part Type']

    # drop all the duplicates across all masked cols - these aren't a problem
    d = power_bi.drop_duplicates(subset=mask)

    # Find duplicated Part Numbers that are left in the dataframe: use duplicated, keep=False
    d2 = d[d.duplicated(subset='Title', keep=False)].sort_values(by='Title')

    dups_out = download_dir / (product + '_duplicates.xlsx')

    write_to_excel(d2[mask], dups_out)

# %%
def addActivate(wb, sheetName):
    try:
        sht = wb.sheets[sheetName].activate()
        print ("sheet activated")
    except ValueError as V:
        print ("Value error sheet didn't exist: {}".format(V))
        sht = wb.sheets.add(sheetName)
        sht = wb.sheets(sheetName).activate()
    except Exception as E:
        sht = wb.sheets.add(sheetName)
        print ("Exception sheet didn't exist: {}".format(V))
        sht = wb.sheets(sheetName).activate()

    return sht

# %%
def mass_roll_up(df):
    g = df[1:][df['has_child'] == 0].groupby(['Function Group','System','Sub System','Level_4_Parent'])['CAD Mass'].agg(['sum','mean','max'])
    import xlwings as xw

    wb = xw.Book()
    ws = wb.sheets[0]

    ws['a1'].options(pd.DataFrame, header=True, index=True).value=g

# %%
def missing_values_table(df):
    mis_val = df.isnull().sum()
    mis_val_percent = 100 * df.isnull().sum() / len(df)
    mis_val_table = pd.concat([mis_val, mis_val_percent], axis=1)
    mis_val_table_ren_columns = mis_val_table.rename(
    columns = {0 : 'Missing Values', 1 : '% of Total Values'})
    mis_val_table_ren_columns = mis_val_table_ren_columns[
        mis_val_table_ren_columns.iloc[:,1] != 0].sort_values(
    '% of Total Values', ascending=False).round(1)
    print ("Your selected dataframe has " + str(df.shape[1]) + " columns.\n"      
        "There are " + str(mis_val_table_ren_columns.shape[0]) +
            " columns that have missing values.")
    return mis_val_table_ren_columns

# %%
def create_heatmap(df, figsize):
    import numpy as np
    import matplotlib.pyplot as plt
    import seaborn as sns

       
    hmap = plt.figure(figsize=figsize)
    ax = sns.heatmap(df, annot = True, fmt=".0%", cmap='YlGnBu', annot_kws={'fontsize':8}, linewidths=0.5)
    ax.set(xlabel="", ylabel="")
    ax.xaxis.tick_top()
    plt.rc('xtick', labelsize=10)
    plt.rc('ytick', labelsize=10)
    cbar = ax.collections[0].colorbar
    cbar.set_ticks([0, .2, .75, 1])
    cbar.set_ticklabels(['0%', '20%', '75%', '100%'])
    plt.figure()
    # sns.set(font_scale=.5)
    # plt.show()
    plt.close(hmap)
    return hmap

# %%
def bar_chart(df, report_fg, figsize):
    import matplotlib.pyplot as plt

    plt.rcParams["figure.figsize"] = figsize
    plt.rcParams["figure.autolayout"] = True

    # fgroups = ['Accessories','Body Exterior','Body Interior','Body Structures','Chassis','Electrical','Powertrain']

    bars = pd.DataFrame()

    for frame in ['P-BoM count','REL P-BoM count','Total Material Lines with PO Coverage']:
        temp_df = []
        # print (reports_dict[frame].reset_index()[reports_dict[frame].reset_index()['Source Code'] == 'Totals'])
        temp_df = df[frame][report_fg].tail(1)
        temp_df = temp_df.rename({'Totals':frame})
        bars = pd.concat([bars, temp_df])
        
    bars = bars.transpose()

    # Creating plot
    ax = bars[['P-BoM count','REL P-BoM count','Total Material Lines with PO Coverage']].plot(kind='bar', title ="Totals", legend=True, fontsize=12)
    # ax.set_xlabel("Hour", fontsize=12)
    ax.set_ylabel("Count of Parts", fontsize=10)
    ax.set_xlabel("")
    plt.rc('xtick', labelsize=10)

    # Call add_value_labels. All the magic happens there.
    add_value_labels(ax)
    ax.set(yticklabels=[])
    for spine in ax.spines:
        ax.spines[spine].set_visible(False)

    # ax = f.add_subplot(1,1,1)
    fig = ax.get_figure()
    plt.close(fig)

    return(fig)
    

# %%
def write_to_xl2(outfile, df_dict):

    import xlwings as xw

    with xw.App(visible=False) as app:
        try:
            wb = xw.Book(outfile)
        except FileNotFoundError:
            wb = xw.Book()
            wb.save(outfile)

        for key in df_dict.keys():
            try:
                ws = wb.sheets.add(key)
            except Exception as e:
                print (e)
            
            ws = wb.sheets[key]

            table_name = key

            ws.clear()

            df = df_dict[key].set_index(list(df_dict[key])[0])
            if table_name in [table.df for table in ws.tables]:
                ws.tables[table_name].update(df)
            else:
                mytable = ws.tables.add(source=ws['A1'],
                                            name=table_name).update(df)

        # # add the BOM sheet
        # try:
        #     ws = wb.sheets.add('BOM')
        # except Exception as e:
        #     print (e)
        
        # ws = wb.sheets['BOM']

        # ws.clear()

        # # write to sheet[0] power_bi df (1st row metrics and BOM COUNT col already removed when creating power_bi df)
        # ws['A1'].options(pd.DataFrame, header=True, index=False).value=bom



        wb.save() 

# %%
def report_metrics(download_dir, product, reports_dict, power_bi):
    """
    Call all data previously written to reports_dict.  Loop through and write out each to the Power bi spreadsheet
    """
    import xlwings as xw
    from openpyxl.utils.cell import get_column_letter

    # if TEST:
    #     report_metrics_filename = os.path.join("TEST_{}_power_bi_metrics.xlsx".format(product))
    # else:
    
    report_metrics_filename = os.path.join(download_dir / "{}_power_bi_metrics.xlsx".format(product))

    with xw.App(visible=False) as app:
        try:
            wb = xw.Book(report_metrics_filename)
            print ("writing to existing {}".format(report_metrics_filename))
        except FileNotFoundError:
            # create a new book
            print ("creating new {}".format(report_metrics_filename))
            wb = xw.Book()
            wb.save(report_metrics_filename)

        # ws = addActivate(wb, 'fg_reports')
        # ws = wb.sheets['fg_reports'].activate()
        
        ws = wb.sheets.add()

        # start with a clean sheet with no contents or formatting
        # ws.clear()
        # ws.autofit(axis="columns")
        report_time = time.strftime(time_format, time.localtime())

        row_offset = 6

        ws['A1'].value = 'Report Time:'
        ws['B1'].value = report_time
        ws['A2'].value = 'BoM last extracted:'
        ws['B2'].value = extract_date

        lightblue=(180,198,231)

        # Process reports_dict

        for report in fg_reports:
            try:
                # logit.info("writing: {}".format(report))
                # color the Header columns
                # find the last column letter
                last_col_letter = get_column_letter(fg_reports[report].shape[1]+2)
                ws['B' + str(row_offset-2)].value=report
                ws['B' + str(row_offset-2)].font.bold=True
                ws['B' + str(row_offset-2)].font.size=16
                ws['B' + str(row_offset)].options(pd.DataFrame, header=1, index=True).value=fg_reports[report]
                ws.range('B{}:{}{}'.format(row_offset, last_col_letter, row_offset)).color=lightblue
                ws.range('B{}:{}{}'.format(row_offset, last_col_letter, row_offset)).font.bold = True
                ws['B' + str(row_offset)].options(pd.DataFrame, header=1, index=True).value=fg_reports[report]
                # this coloured the last line in the table lightblue (for when there are totals)
                # ws.range('B{}:{}{}'.format(row_offset + reports_dict[report].shape[0], last_col_letter, row_offset + reports_dict[report].shape[0])).color=lightblue
                # this adds bold to the last line in the table (for when there are totals)
                # ws.range('B{}:{}{}'.format(row_offset + fg_reports[report].shape[0], last_col_letter, row_offset + fg_reports[report].shape[0])).font.bold = True
                row_offset = row_offset + fg_reports[report].shape[0] + 7
            except AttributeError:
                # probably writing out an image rather than a dataframe
                ws['B' + str(row_offset-2)].value=report
                ws['B' + str(row_offset-2)].font.bold=True
                ws['B' + str(row_offset-2)].font.size=16            
                ws.pictures.add(fg_reports[report], name=report, update=True, left=ws.range('B' + str(row_offset)).left, top=ws.range('B' + str(row_offset)).top)
                row_offset = row_offset + 32
            except Exception as err:
                # logit.exception(f"Unexpected {err=}, {type(err)=}")
                print(f"Unexpected {err=}, {type(err)=}")
                raise

            # outrow += reports_dict[report].shape[0]+7
            # ws.pictures.add(hmap, name="REL % vs P-BoM Count", update=True, left=ws.range('M' + str(outrow)).left, top=ws.range('M' + str(outrow)).top)
            # outrow += 17

            # ws = addActivate(wb, 'fg_combined')
            # ws = wb.sheets['fg_combined'].activate()
            ws = wb.sheets.add()

            # start with a clean sheet with no contents or formatting
            # ws.clear()
            # ws.autofit(axis="columns")
            report_time = time.strftime(time_format, time.localtime())

            row_offset = 6

            ws['A1'].value = 'Report Time:'
            ws['B1'].value = report_time
            ws['A2'].value = 'BoM last extracted:'
            ws['B2'].value = extract_date

            lightblue=(180,198,231)

            # Process reports_dict

            for report in fg_combined:
                try:
                    # logit.info("writing: {}".format(report))
                    # color the Header columns
                    # find the last column letter
                    last_col_letter = get_column_letter(fg_combined[report].shape[1]+2)
                    ws['B' + str(row_offset-2)].value=report
                    ws['B' + str(row_offset-2)].font.bold=True
                    ws['B' + str(row_offset-2)].font.size=16
                    ws['B' + str(row_offset)].options(pd.DataFrame, header=1, index=True).value=fg_combined[report]
                    ws.range('B{}:{}{}'.format(row_offset, last_col_letter, row_offset)).color=lightblue
                    ws.range('B{}:{}{}'.format(row_offset, last_col_letter, row_offset)).font.bold = True
                    ws['B' + str(row_offset)].options(pd.DataFrame, header=1, index=True).value=fg_combined[report]
                    # ws.range('B{}:{}{}'.format(row_offset + reports_dict[report].shape[0], last_col_letter, row_offset + reports_dict[report].shape[0])).color=lightblue
                    ws.range('B{}:{}{}'.format(row_offset + fg_combined[report].shape[0], last_col_letter, row_offset + fg_combined[report].shape[0])).font.bold = True
                    row_offset = row_offset + fg_combined[report].shape[0] + 7
                except AttributeError:
                    # probably writing out an image rather than a dataframe
                    ws['B' + str(row_offset-2)].value=report
                    ws['B' + str(row_offset-2)].font.bold=True
                    ws['B' + str(row_offset-2)].font.size=16            
                    ws.pictures.add(fg_combined[report], name=report, update=True, left=ws.range('B' + str(row_offset)).left, top=ws.range('B' + str(row_offset)).top)
                    row_offset = row_offset + 32
                except Exception as err:
                    # logit.exception(f"Unexpected {err=}, {type(err)=}")
                    print(f"Unexpected {err=}, {type(err)=}")
                    raise

            # ws = addActivate(wb, 'BOM')
            # ws = wb.sheets['BOM'].activate()
            ws = wb.sheets.add()
            # start with a clean sheet with no contents or formatting
            # ws.clear()
            # ws.autofit(axis="columns")
            report_time = time.strftime(time_format, time.localtime())

            # write to sheet[0] power_bi df (1st row metrics and BOM COUNT col already removed when creating power_bi df)
            ws['A1'].options(pd.DataFrame, header=True, index=False).value=power_bi

            wb.save()


# %% [markdown]
# # Main Processing

# %%
if __name__ == '__main__':

    if 'macOS' in platform.platform():
        # set some defaults for testing on mac
        download_dir = Path('/Users/mark/Downloads')
        user_dir = download_dir
        sharepoint_dir = download_dir

    elif 'Server' in platform.platform():
        # we're on the azure server (probably)
        user_dir = Path('Z:/python/FilesIn')

        download_dir = Path(user_dir)
        user_dir = download_dir
        sharepoint_dir = Path('Z:/python/FilesOut')

    elif os.getlogin() == 'mark_':
        # my test windows machine
        download_dir = Path('C:/Users/mark_/Downloads')
        user_dir = download_dir
        sharepoint_dir = download_dir        

    else:
        # personal one drive
        user_dir = 'C:/Users/USERNAME'

        # replace USERNAME with current logged on user
        user_dir = user_dir.replace('USERNAME', os.getlogin())

        # read in config file
        config = configparser.ConfigParser()
        config.read('user_directory.ini')

        # read in gm_dir and gm_docs from config file
        gm_dir = Path(config[os.getlogin().lower()]['gm_dir'])
        gm_docs = Path(config[os.getlogin().lower()]['gmt'])
        # this may find more than one sharepoint directory
        # sharepoint_dir = user_dir + "/" + gm_dir + "/" + gm_docs
        sharepoint_dir = Path(user_dir / gm_dir / gm_docs)

        # download_dir = os.path.join(sharepoint_dir, 'Data Shuttle', 'downloads')
        download_dir = Path(sharepoint_dir / 'Data Shuttle' / 'downloads')

    files = find_files(download_dir)

    if len(files) == 0:
        print ("No files found in {}".format(download_dir))
       

    else:
        dict_df = {}

        # loop in case more than 1 file
        for file in files:
            BOM = pd.DataFrame()
            BOM = open_file(file)
            orig_bom = BOM.copy()
            
            time_format = "%Y-%m-%d %H:%M"
            curr_time = time.strftime(time_format, time.localtime())

            # check mandatory attributes are present
            mandatory_cols = mandatory_attributes(BOM.columns)

            # if 'Function Group' not in BOM.columns:
            BOM = add_function_group(BOM)
            BOM = rename_columns(BOM)

            # read the title from first row as product to this file after
            product = BOM['Title'].loc[0]

            # populate COG x, y, z from COG field
            # 04/04/2024 - Jannik/Carlos - don't need separate COG cols anymore
            # BOM = COG_split(BOM)

            # add an SA_Index for the add_quantities stage
            BOM_sa = create_sa_index(BOM)

            BOM_pp = create_parent_part(BOM_sa)
            #  don't want full MK after all!
            # BOM_pp = create_gparent_part(BOM_sa)

            # if we've not been given quantity we need to do the roll-up ourselves
            if 'Quantity' not in BOM.columns:
                # this creates the quantity column
                BOM_pp = add_quantities(BOM_pp)

            # don't include 'Part Number' from 3dx - it's not the real part number and confuses Carlos' process
            try:
                BOM_pp.drop('Part Number', axis=1, inplace=True)
            except:
                pass
            # don't keep SA_Index in output as not needed.            
            try:
                BOM_pp.drop('SA_Index', axis=1, inplace=True)
            except:
                pass

            # add part number length
            BOM_pp['part_number_length'] = BOM_pp['Title'].str.len()
            
            # BOM_ordered['Effectivity'] = BOM_ordered['Effectivity'].apply(convert_date)

            dict_df[product] = BOM_pp

            # replace only spaces in cells with NaN
            BOM_pp = clear_blanks(BOM_pp)
            # replace zero values with NaN
            BOM_pp = clear_zero_values(BOM_pp)
            # replace 'Not Set' values with NaN
            BOM_pp = clear_not_set_values(BOM_pp)
            # populate SYS and ENG source codes with 'N/A'
            BOM_pp = set_na_values(BOM_pp)

            # BOM_pp = split_subtype(BOM_pp)

            # write out the updated filename with timestamp to the correct dir
            project = product.split('-')[0]
            output_file = 'Updated_{}_{}.xlsx'.format(product, curr_time.split()[0])
            output_path = os.path.join(sharepoint_dir, project, output_file)
            # automatically create the parent directories if don't exist
            Path(output_path).parent.mkdir(parents=True, exist_ok=True)            
            # print ("file written to {}".format(output_path))

            # drop the filename timestamp for power bi 
            # power_bi_file = 'Updated_{}.xlsx'.format(product)
            # write to data shuttle directory for Carlos to pick up            
            data_shuttle_file = 'Updated_{}_{}.xlsx'.format(product, curr_time.split()[0])
            ds_folder = data_shuttle_folder(product)
            data_shuttle_path = os.path.join(sharepoint_dir, 'Data Shuttle', ds_folder, data_shuttle_file)
            # write out without timestamp for power bi
            # power_bi_path = os.path.join(sharepoint_dir, 'Data Shuttle', ds_folder, power_bi_file)
            # automatically create the parent directories if don't exist
            Path(data_shuttle_path).parent.mkdir(parents=True, exist_ok=True) 

            BOM_pp = add_matching_key(BOM_pp)

            # st_birthtime = creation time on MacOS, needs st_ctime for Windows 
            if 'macOS' in platform.platform():
                extract_date = datetime.datetime.fromtimestamp(Path(file).stat().st_birthtime)
            else:
                extract_date = datetime.datetime.fromtimestamp(Path(file).stat().st_ctime)
                
            BOM_pp = add_smartsheet_cols(BOM_pp, extract_date)            
            # format Last Export Date as datetime, with dayfirst
            BOM_pp['Last Export Date'] = pd.to_datetime(BOM_pp['Last Export Date'], dayfirst=True)
            # sort by orig_sort before writing out
            BOM_pp = BOM_pp.sort_values(by='orig_sort')

            BOM_pp = has_child(BOM_pp)

            # drop packaging function group from the extract we send to data shuttle for processing
            BOM_without_packaging = BOM_pp[~BOM_pp['Function Group'].str.contains('PACKAGING', na=False)]

            # calculate percent missing after dropping packaging
            BOM_without_packaging = percent_missing(BOM_without_packaging)
            BOM_pp = percent_missing(BOM_pp)
            BOM_without_packaging.loc['percent_missing','BOM COUNT'] = BOM_without_packaging.shape[0]
            BOM_pp.loc['percent_missing','BOM COUNT'] = BOM_pp.shape[0]

            # order the cols
            BOM_ordered_without_packaging = order_columns(BOM_without_packaging)
            BOM_ordered = order_columns(BOM_pp)
            # write out the full file
            write_to_excel(BOM_ordered, output_path)
            # write out without packaging to data shuttle
            write_to_excel(BOM_ordered_without_packaging, data_shuttle_path)
            print ("file written to {}".format(data_shuttle_path))

            # power_bi processing
            power_bi = create_power_bi_df(BOM_ordered_without_packaging, product)
            # add part validation
            power_bi = validate_part_no(power_bi)
            reports_dict = build_reports(power_bi)
            reports_dict['BOM'] = power_bi
            # report_metrics(download_dir, product, reports_dict, power_bi)
            # this would have written to a power_bi folder.  Let's leave that for time being
            power_bi_outfile = sharepoint_dir / "power_bi" / "{}_power_bi_metrics.xlsx".format(product)
            # out_file = os.path.join(output_dir / "{}_power_bi_metrics.xlsx".format(product))
            Path(power_bi_outfile).parent.mkdir(parents=True, exist_ok=True)            
            write_to_xl2(power_bi_outfile, reports_dict)

            


# %%



